"""Loader metadanych osób z pliku '10 osoby - ok.xlsm'.

Plik ma strukturę:
- Arkusz 'Dane' — słowniki referencyjne (gender, skin_tone, unique_characteristics).
- Arkusze '01'..'22' — po jednym na osobę. Wiersz 4 zawiera dane osobowe
  (gender, age, skin_tone, unique_characteristics, facial_hair_interference,
   lipstick, lip_balm, excessive_reflections).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl


# Indeksy kolumn w wierszu 4 arkuszy osób (1-indexed jak w openpyxl).
_COL_GENDER = 1
_COL_AGE = 2
_COL_SKIN_TONE = 3
_COL_UNIQUE = 4
_COL_FACIAL_HAIR = 5
_COL_LIPSTICK = 6
_COL_LIP_BALM = 7
_COL_REFLECTIONS = 8

_DATA_ROW = 4  # wiersz z wartościami w arkuszu osoby


@dataclass
class PersonMetadata:
    """Metadane jednej osoby z pliku Excel."""

    person_id: int
    gender: Optional[str] = None  # "M" / "F"
    age: Optional[int] = None
    skin_tone: Optional[str] = None  # "light" / "medium" / "dark"
    unique_characteristics: Optional[str] = None  # np. "none", "injury_cuts", "freckles"
    facial_hair_interference: Optional[str] = None  # "Yes" / "No"
    lipstick: Optional[str] = None  # "Yes" / "No"
    lip_balm: Optional[str] = None  # "Yes" / "No"
    excessive_reflections: Optional[str] = None  # "Yes" / "No"

    def to_dict(self) -> Dict[str, object]:
        """Zwraca metadane jako słownik (np. do wyświetlenia w GUI)."""
        return {
            "person_id": self.person_id,
            "gender": self.gender,
            "age": self.age,
            "skin_tone": self.skin_tone,
            "unique_characteristics": self.unique_characteristics,
            "facial_hair_interference": self.facial_hair_interference,
            "lipstick": self.lipstick,
            "lip_balm": self.lip_balm,
            "excessive_reflections": self.excessive_reflections,
        }


@dataclass
class MetadataDictionaries:
    """Słowniki referencyjne z arkusza 'Dane'."""

    genders: List[str] = field(default_factory=list)
    skin_tones: List[str] = field(default_factory=list)
    unique_characteristics: Dict[str, str] = field(default_factory=dict)  # EN → PL


def _safe_str(value) -> Optional[str]:
    """Konwertuje wartość komórki na string, zwraca None dla pustych."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _safe_int(value) -> Optional[int]:
    """Konwertuje wartość komórki na int, zwraca None dla pustych/błędnych."""
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def load_person_metadata(xlsm_path: Path | str) -> Dict[int, PersonMetadata]:
    """Wczytaj metadane wszystkich osób z pliku xlsm.

    Args:
        xlsm_path: ścieżka do pliku '10 osoby - ok.xlsm'.

    Returns:
        Słownik person_id (int 1-22) → PersonMetadata.
    """
    xlsm_path = Path(xlsm_path)
    if not xlsm_path.exists():
        raise FileNotFoundError(f"Plik metadanych nie istnieje: {xlsm_path}")

    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=False)

    result: Dict[int, PersonMetadata] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "Dane":
            continue
        try:
            person_id = int(sheet_name)
        except ValueError:
            continue

        ws = wb[sheet_name]
        result[person_id] = PersonMetadata(
            person_id=person_id,
            gender=_safe_str(ws.cell(row=_DATA_ROW, column=_COL_GENDER).value),
            age=_safe_int(ws.cell(row=_DATA_ROW, column=_COL_AGE).value),
            skin_tone=_safe_str(ws.cell(row=_DATA_ROW, column=_COL_SKIN_TONE).value),
            unique_characteristics=_safe_str(ws.cell(row=_DATA_ROW, column=_COL_UNIQUE).value),
            facial_hair_interference=_safe_str(ws.cell(row=_DATA_ROW, column=_COL_FACIAL_HAIR).value),
            lipstick=_safe_str(ws.cell(row=_DATA_ROW, column=_COL_LIPSTICK).value),
            lip_balm=_safe_str(ws.cell(row=_DATA_ROW, column=_COL_LIP_BALM).value),
            excessive_reflections=_safe_str(ws.cell(row=_DATA_ROW, column=_COL_REFLECTIONS).value),
        )

    wb.close()
    return result


def load_dictionaries(xlsm_path: Path | str) -> MetadataDictionaries:
    """Wczytaj słowniki referencyjne z arkusza 'Dane'.

    Słowniki zawierają możliwe wartości pól (gender, skin_tone, unique_characteristics)
    oraz tłumaczenia EN → PL dla unique_characteristics.
    """
    xlsm_path = Path(xlsm_path)
    if not xlsm_path.exists():
        raise FileNotFoundError(f"Plik metadanych nie istnieje: {xlsm_path}")

    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=False)
    if "Dane" not in wb.sheetnames:
        wb.close()
        return MetadataDictionaries()

    ws = wb["Dane"]
    dicts = MetadataDictionaries()

    section: Optional[str] = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        first = _safe_str(row[0])
        if first is None:
            section = None
            continue

        if first == "gender":
            section = "gender"
            continue
        if first == "skin_tone":
            section = "skin_tone"
            continue
        if first == "unique_characteristics":
            section = "unique_characteristics"
            continue
        if first == "facial_hair_interference / lipstick / lib_balm / escessive_reflections":
            section = None
            continue

        if section == "gender":
            dicts.genders.append(first)
        elif section == "skin_tone":
            dicts.skin_tones.append(first)
        elif section == "unique_characteristics":
            pl_name = _safe_str(row[1]) if len(row) > 1 else None
            dicts.unique_characteristics[first] = pl_name or first

    wb.close()
    return dicts
