"""Zapis wyników eksperymentów do pliku Excel.

Arkusze (ExperimentResults):
    Run_info        — parametry uruchomienia, timestamp, hash
    Predictions     — każda predykcja: plik, true, predicted, top3/5, czas
    Summary_M2_<X>  — podsumowanie Metody 2 per klasyfikator
    ConfusionMatrix_<method> — macierz 22×22
    FeatureStats    — statystyki cech minucji per osoba

Funkcja write_detailed_prediction_excel():
    Zapisuje/dopisuje jeden wiersz predykcji do pliku exports/predykcje_szczegolowe.xlsx
    z pełnymi danymi o preprocessingu, cechach i wynikach wszystkich 3 klasyfikatorów.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.config import EXPORTS_DIR, RESULTS_DIR


# ---------------------------------------------------------------------------
# Legenda kolumn dla szczegółowego arkusza predykcji
# Każdy wpis: (nazwa_kolumny, opis, jednostka_lub_zakres)
# ---------------------------------------------------------------------------
_DETAILED_COLUMNS: List[Tuple[str, str, str]] = [
    # META
    ("plik",                      "Nazwa pliku zdjęcia ust",                             "np. 2026-01-06_..._Nr_01_LampaON_..._CUT.png"),
    ("osoba_prawdziwa",            "Numer osoby wyekstrahowany z nazwy pliku",             "Nr_01 ... Nr_22  (lub 'zewnetrzne')"),
    ("lampa",                     "Czy lampa doświetlająca była włączona",                "ON / OFF / brak_info"),
    ("data_zdjecia",               "Data wykonania zdjęcia z nazwy pliku",                 "YYYY-MM-DD"),
    ("typ_zdjecia",                "Źródło zdjęcia",                                      "z_bazy / zewnetrzne"),
    # PREPROCESSING
    ("oryg_szerokosc_px",          "Szerokość oryginału w pikselach",                     "piksele"),
    ("oryg_wysokosc_px",           "Wysokość oryginału w pikselach",                      "piksele"),
    ("oryg_srednia_jasnosc",       "Średnia jasności pikseli oryginału (0–255)",           "0–255"),
    ("oryg_odch_std",              "Odchylenie standardowe jasności oryginału",            "0–128"),
    ("clahe_clip_limit",           "Parametr clipLimit CLAHE (ogr. kontrastu)",           "typowo 2.0"),
    ("clahe_rozmiar_kafla",        "Rozmiar kafla CLAHE (tileGridSize × tileGridSize)",   "piksele, typowo 8"),
    ("bilateral_d",                "Średnica sąsiedztwa bilateral filter",                "piksele, typowo 9"),
    ("bilateral_sigma_kolor",      "Sigma w przestrzeni kolorów (bilateral)",             "typowo 75"),
    ("bilateral_sigma_przestrzen", "Sigma geometryczna bilateral filter",                 "typowo 75"),
    ("final_szerokosc_px",         "Szerokość obrazu po resize (stała)",                  "512 px"),
    ("final_wysokosc_px",          "Wysokość obrazu po resize (stała)",                   "256 px"),
    ("final_srednia_jasnosc",      "Średnia jasności po pełnym preprocessingu",           "0–255"),
    ("final_odch_std",             "Odchylenie std. jasności po preprocessingu",          "0–128"),
    # CECHY LBP (Local Binary Patterns)
    ("lbp_bin_0",  "LBP bin 0 — udział wzorca '00000000' (jednolity)",  "0–1, znormalizowany"),
    ("lbp_bin_1",  "LBP bin 1",                                          "0–1"),
    ("lbp_bin_2",  "LBP bin 2",                                          "0–1"),
    ("lbp_bin_3",  "LBP bin 3",                                          "0–1"),
    ("lbp_bin_4",  "LBP bin 4",                                          "0–1"),
    ("lbp_bin_5",  "LBP bin 5",                                          "0–1"),
    ("lbp_bin_6",  "LBP bin 6",                                          "0–1"),
    ("lbp_bin_7",  "LBP bin 7",                                          "0–1"),
    ("lbp_bin_8",  "LBP bin 8",                                          "0–1"),
    ("lbp_bin_9",  "LBP bin 9 — udział wzorca '11111111' (jednolity)",  "0–1"),
    ("lbp_entropia",               "Entropia histogramu LBP — różnorodność tekstury",    "bitów (0–3.32 dla 10 binów)"),
    # CECHY HOG (Histogram of Oriented Gradients)
    ("hog_energia",                "Średnia kwadratów wartości HOG — ogólna siła krawędzi", "wartość dodatnia"),
    ("hog_orientacja_dominujaca",  "Kąt dominującego gradientu (bin z max sumą HOG)",   "0–180 stopni"),
    # CECHY GABOR
    ("gabor_srednia",              "Średnia odpowiedź Gabora (8 orientacji × 4 częst.)", "wartość float"),
    ("gabor_max",                  "Maksymalna odpowiedź filtra Gabora",                 "wartość float"),
    ("gabor_orientacja_dominujaca","Orientacja (stopnie) filtra z max odpowiedzią Gabora","0–157.5 stopni (co 22.5°)"),
    # CECHY MINUCJI
    ("min_zakonczenia",            "Liczba zakończeń bruzd (CN=1) — miejsca gdzie bruzda się kończy",  "liczba całk. ≥ 0"),
    ("min_rozwidlenia",            "Liczba rozwidleń (CN=3) — miejsca podziału bruzdy",               "liczba całk. ≥ 0"),
    ("min_skrzyzowania",           "Liczba skrzyżowań (CN≥4) — miejsca przecięcia bruzd",             "liczba całk. ≥ 0"),
    ("min_total",                  "Łączna liczba minucji (zakończenia + rozwidlenia + skrzyżowania)", "liczba całk. ≥ 0"),
    ("gestosc_bruzd",              "Stosunek min_total do pikseli szkieletu — zagęszczenie punktów",   "0–1 (float)"),
    ("piksele_szkieletu",          "Liczba pikseli bruzd po binaryzacji i szkieletyzacji",             "piksele ≥ 0"),
    ("entropia_binarna",           "Entropia obrazu binarnego bruzd — złożoność wzorca",               "0–1 (float)"),
    # WYNIKI SVM
    ("svm_top1",       "SVM — przewidywana osoba (top-1)",                               "Nr_01 ... Nr_22"),
    ("svm_conf",       "SVM — pewność (prawdopodobieństwo Platt scaling) top-1",         "0–1 (im wyższe tym pewniejsze)"),
    ("svm_top2",       "SVM — drugi kandydat",                                            "Nr_XX"),
    ("svm_conf2",      "SVM — pewność drugiego kandydata",                                "0–1"),
    ("svm_top3",       "SVM — trzeci kandydat",                                           "Nr_XX"),
    ("svm_conf3",      "SVM — pewność trzeciego kandydata",                               "0–1"),
    ("svm_top4",       "SVM — czwarty kandydat",                                          "Nr_XX"),
    ("svm_conf4",      "SVM — pewność czwartego kandydata",                               "0–1"),
    ("svm_top5",       "SVM — piąty kandydat",                                            "Nr_XX"),
    ("svm_conf5",      "SVM — pewność piątego kandydata",                                 "0–1"),
    ("svm_poprawny",   "Czy SVM trafnie zidentyfikował osobę",                            "TAK / NIE / brak_etykiety"),
    # WYNIKI RF
    ("rf_top1",        "Random Forest — przewidywana osoba (top-1)",                     "Nr_01 ... Nr_22"),
    ("rf_conf",        "RF — pewność top-1 (głosowanie drzew z wagą)",                   "0–1"),
    ("rf_top2",        "RF — drugi kandydat",                                             "Nr_XX"),
    ("rf_conf2",       "RF — pewność drugiego kandydata",                                 "0–1"),
    ("rf_top3",        "RF — trzeci kandydat",                                            "Nr_XX"),
    ("rf_conf3",       "RF — pewność trzeciego kandydata",                                "0–1"),
    ("rf_top4",        "RF — czwarty kandydat",                                           "Nr_XX"),
    ("rf_conf4",       "RF — pewność czwartego kandydata",                                "0–1"),
    ("rf_top5",        "RF — piąty kandydat",                                             "Nr_XX"),
    ("rf_conf5",       "RF — pewność piątego kandydata",                                  "0–1"),
    ("rf_poprawny",    "Czy RF trafnie zidentyfikował osobę",                             "TAK / NIE / brak_etykiety"),
    # WYNIKI k-NN
    ("knn_top1",       "k-NN — przewidywana osoba (top-1)",                              "Nr_01 ... Nr_22"),
    ("knn_conf",       "k-NN — pewność top-1 (softmax z odległości do klas)",            "0–1"),
    ("knn_top2",       "k-NN — drugi kandydat",                                           "Nr_XX"),
    ("knn_conf2",      "k-NN — pewność drugiego kandydata",                               "0–1"),
    ("knn_top3",       "k-NN — trzeci kandydat",                                          "Nr_XX"),
    ("knn_conf3",      "k-NN — pewność trzeciego kandydata",                              "0–1"),
    ("knn_top4",       "k-NN — czwarty kandydat",                                         "Nr_XX"),
    ("knn_conf4",      "k-NN — pewność czwartego kandydata",                              "0–1"),
    ("knn_top5",       "k-NN — piąty kandydat",                                           "Nr_XX"),
    ("knn_conf5",      "k-NN — pewność piątego kandydata",                                "0–1"),
    ("knn_poprawny",   "Czy k-NN trafnie zidentyfikował osobę",                          "TAK / NIE / brak_etykiety"),
    # PODSUMOWANIE
    ("ile_clf_zgodnych", "Ile klasyfikatorów wskazało tę samą osobę co top-1 SVM",      "0 / 1 / 2 / 3"),
    ("konsensus_top1",   "Osoba wskazana przez ≥2 klasyfikatory (lub '-' gdy brak)",    "Nr_XX lub '-'"),
    ("czas_total_ms",    "Łączny czas przetwarzania (preprocessing + cechy + predykcja)","milisekundy"),
    ("timestamp",        "Data i godzina zapisu tego wiersza",                           "YYYY-MM-DD HH:MM:SS"),
]


def write_detailed_prediction_excel(
    *,
    image_name: str,
    true_id: Optional[int],
    is_external: bool,
    prep_result,           # PreprocessingResult
    raw_features: dict,    # {'lbp': np.ndarray, 'hog': ..., 'gabor': ..., 'minutiae': ...}
    predictions: dict,     # {'svm': {...}, 'rf': {...}, 'knn': {...}}  — wynik predict_single
    time_total_ms: float,
    output_path: Optional[Path] = None,
) -> Path:
    """Zapisz/dopisz jeden wiersz predykcji do bogatego pliku Excel.

    Jeśli plik nie istnieje — tworzy go z arkuszem Legenda i nagłówkami.
    Kolejne wywołania dopisują wiersze do arkusza Predykcja.

    Args:
        image_name:    nazwa pliku (lub "upload_<timestamp>").
        true_id:       prawdziwa osoba (None = nieznana / external).
        is_external:   True gdy zdjęcie spoza bazy.
        prep_result:   PreprocessingResult z src.method2_ml.preprocessing.
        raw_features:  słownik z extract_raw_features() {'lbp', 'hog', 'gabor', 'minutiae'}.
        predictions:   {'svm': predict_single_result, 'rf': ..., 'knn': ...}.
        time_total_ms: czas przetwarzania w ms.
        output_path:   ścieżka docelowa (domyślnie exports/predykcje_szczegolowe.xlsx).

    Returns:
        Ścieżka do pliku Excel.
    """
    import re

    if output_path is None:
        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        output_path = EXPORTS_DIR / "predykcje_szczegolowe.xlsx"
    output_path = Path(output_path)

    # Wczytaj lub stwórz workbook
    if output_path.exists():
        wb = openpyxl.load_workbook(str(output_path))
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _write_legenda_sheet(wb)
        _write_prediction_headers(wb)

    ws = wb["Predykcja"]

    # --- Parsuj nazwę pliku ---
    m = re.search(r'(\d{4}-\d{2}-\d{2}).*Nr_?(\d+).*(LampaON|LampaOFF)', image_name)
    date_str  = m.group(1) if m else "brak"
    lamp_str  = m.group(3) if m else "brak_info"
    if true_id is not None:
        osoba_str = f"Nr_{true_id:02d}"
    elif not is_external and m:
        osoba_str = f"Nr_{int(m.group(2)):02d}"
    else:
        osoba_str = "zewnetrzne"

    # --- Statystyki preprocessingu ---
    orig_h, orig_w = prep_result.original_bgr.shape[:2]
    orig_gray = prep_result.grayscale
    orig_mean = float(np.mean(orig_gray))
    orig_std  = float(np.std(orig_gray))
    final     = prep_result.resized
    fin_mean  = float(np.mean(final))
    fin_std   = float(np.std(final))

    # --- Statystyki LBP ---
    lbp = raw_features["lbp"].astype(np.float64)
    lbp_norm = lbp / (lbp.sum() + 1e-10)
    lbp_entropy = float(-np.sum(lbp_norm * np.log2(lbp_norm + 1e-12)))

    # --- Statystyki HOG ---
    hog_vec = raw_features["hog"].astype(np.float64)
    hog_energy = float(np.mean(hog_vec ** 2))
    # HOG bins: 9 orientations, sum across all cells → dominant angle
    n_orient = 9
    bins_per_orient = hog_energy  # placeholder
    # Reshape: (cells, orient) approx — sum bins that are multiples of 9
    hog_orient_hist = np.zeros(n_orient)
    for k in range(n_orient):
        hog_orient_hist[k] = float(np.sum(hog_vec[k::n_orient]))
    dom_orient_idx = int(np.argmax(hog_orient_hist))
    hog_dom_angle = dom_orient_idx * (180.0 / n_orient)

    # --- Statystyki Gabor ---
    gabor_vec = raw_features["gabor"].astype(np.float64)
    gabor_mean = float(np.mean(gabor_vec))
    gabor_max  = float(np.max(gabor_vec))
    # 8 orientacje (co 22.5°), wartości są zorganizowane jako (orient0_f0, orient0_f1, ..., orient7_f3)
    n_orient_g = 8
    gabor_per_orient = [float(np.mean(gabor_vec[i::n_orient_g])) for i in range(n_orient_g)]
    dom_orient_g = int(np.argmax(gabor_per_orient))
    gabor_dom_angle = dom_orient_g * (180.0 / n_orient_g)

    # --- Minucje ---
    min_vec = raw_features["minutiae"]
    n_end  = int(round(float(min_vec[0])))
    n_bif  = int(round(float(min_vec[1])))
    n_cross= int(round(float(min_vec[2])))
    n_tot  = int(round(float(min_vec[3])))
    groove_dens = float(min_vec[4])
    skel_px     = int(round(float(min_vec[5])))
    bin_entropy = float(min_vec[6])

    # --- Wyniki klasyfikatorów ---
    def _fmt_pred(pred_dict, k):
        top_k = pred_dict.get("top_k", [])
        if k - 1 < len(top_k):
            pid, conf = top_k[k - 1]
            return f"Nr_{pid:02d}", round(conf, 4)
        return "-", 0.0

    def _correct(pred_dict, true):
        if true is None:
            return "brak_etykiety"
        return "TAK" if pred_dict.get("predicted") == true else "NIE"

    svm = predictions["svm"]
    rf  = predictions["rf"]
    knn = predictions["knn"]

    # Konsensus
    tops = [svm["predicted"], rf["predicted"], knn["predicted"]]
    from collections import Counter
    cnt = Counter(tops)
    most_common_id, most_common_n = cnt.most_common(1)[0]
    konsensus = f"Nr_{most_common_id:02d}" if most_common_n >= 2 else "-"
    ile_zgodnych = sum(1 for t in tops if t == svm["predicted"])

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    row = [
        # META
        image_name, osoba_str, lamp_str, date_str,
        "zewnetrzne" if is_external else "z_bazy",
        # PREPROCESSING
        orig_w, orig_h, round(orig_mean, 2), round(orig_std, 2),
        prep_result.clahe_clip_limit, prep_result.clahe_tile_size,
        prep_result.bilateral_d, prep_result.bilateral_sigma_color, prep_result.bilateral_sigma_space,
        final.shape[1], final.shape[0], round(fin_mean, 2), round(fin_std, 2),
        # LBP
        *[round(float(v), 6) for v in lbp_norm],
        round(lbp_entropy, 4),
        # HOG
        round(hog_energy, 6), round(hog_dom_angle, 1),
        # Gabor
        round(gabor_mean, 4), round(gabor_max, 4), round(gabor_dom_angle, 1),
        # Minucje
        n_end, n_bif, n_cross, n_tot, round(groove_dens, 6), skel_px, round(bin_entropy, 4),
        # SVM
        f"Nr_{svm['predicted']:02d}", round(svm["confidence"], 4),
        *[x for k in range(2, 6) for x in _fmt_pred(svm, k)],
        _correct(svm, true_id),
        # RF
        f"Nr_{rf['predicted']:02d}", round(rf["confidence"], 4),
        *[x for k in range(2, 6) for x in _fmt_pred(rf, k)],
        _correct(rf, true_id),
        # k-NN
        f"Nr_{knn['predicted']:02d}", round(knn["confidence"], 4),
        *[x for k in range(2, 6) for x in _fmt_pred(knn, k)],
        _correct(knn, true_id),
        # PODSUMOWANIE
        ile_zgodnych, konsensus, round(time_total_ms, 1), ts,
    ]

    ws.append(row)

    # Pokoloruj wiersz jeśli poprawne
    last_row = ws.max_row
    if true_id is not None:
        for col_idx, col_name in enumerate([c[0] for c in _DETAILED_COLUMNS], start=1):
            if col_name in ("svm_poprawny", "rf_poprawny", "knn_poprawny"):
                cell = ws.cell(row=last_row, column=col_idx)
                if cell.value == "TAK":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif cell.value == "NIE":
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")

    wb.save(str(output_path))
    return output_path


def _write_legenda_sheet(wb: openpyxl.Workbook) -> None:
    """Utwórz arkusz Legenda z opisem każdej kolumny."""
    ws = wb.create_sheet("Legenda")

    # Tytuł
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = "Legenda — opis kolumn arkusza 'Predykcja'"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="2E75B6")
    title_cell.alignment = Alignment(horizontal="center")

    ws.append([])  # pusty wiersz

    headers = ["Nazwa kolumny", "Opis", "Jednostka / możliwe wartości"]
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Grupy kolumn — nagłówki sekcji
    _groups = {
        "plik":           ("META — informacje o zdjęciu",           "70FF70"),
        "oryg_szerokosc_px": ("PREPROCESSING — parametry i statystyki obrazu", "FFE699"),
        "lbp_bin_0":      ("CECHY LBP — Local Binary Patterns (tekstura)",    "FCE4D6"),
        "hog_energia":    ("CECHY HOG — Histogram of Oriented Gradients (krawędzie)", "DDEBF7"),
        "gabor_srednia":  ("CECHY GABOR — odpowiedzi filtrów Gabora (bruzdy kierunkowe)", "EAF2D3"),
        "min_zakonczenia":("CECHY MINUCJI — punkty charakterystyczne bruzd",  "F2CEEF"),
        "svm_top1":       ("WYNIKI SVM (Support Vector Machine)",              "FFDAB9"),
        "rf_top1":        ("WYNIKI RF (Random Forest)",                        "B0E0E6"),
        "knn_top1":       ("WYNIKI k-NN (k Nearest Neighbours)",               "FFFACD"),
        "ile_clf_zgodnych":("PODSUMOWANIE — zgodność klasyfikatorów",          "E2EFDA"),
    }

    current_group = ""
    for col_name, desc, unit in _DETAILED_COLUMNS:
        if col_name in _groups:
            current_group, grp_color = _groups[col_name]
            # Sekcja
            ws.append([])
            row_idx = ws.max_row
            ws.cell(row=row_idx, column=1, value=f"── {current_group} ──")
            ws.cell(row=row_idx, column=1).font = Font(bold=True, italic=True)
            ws.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=grp_color)
            ws.merge_cells(f"A{row_idx}:C{row_idx}")

        row_data = [col_name, desc, unit]
        ws.append(row_data)
        r = ws.max_row
        ws.cell(r, 1).font = Font(bold=True, name="Consolas")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 45
    ws.row_dimensions[1].height = 22


def _write_prediction_headers(wb: openpyxl.Workbook) -> None:
    """Utwórz arkusz Predykcja z nagłówkami."""
    ws = wb.create_sheet("Predykcja")
    headers = [c[0] for c in _DETAILED_COLUMNS]
    ws.append(headers)
    for i, cell in enumerate(ws[1]):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    # Szerokości kolumn
    for col_name, _, _ in _DETAILED_COLUMNS:
        idx = [c[0] for c in _DETAILED_COLUMNS].index(col_name) + 1
        ws.column_dimensions[get_column_letter(idx)].width = 18
    ws.column_dimensions["A"].width = 55  # plik


@dataclass
class PredictionRow:
    """Jeden wiersz tabeli predykcji."""

    image_path: str
    true_id: int
    method: str          # "ssim", "orb", "hist", "combined", "svm", "rf", "knn"
    predicted_id: int
    top3: List[int]      # lista top-3 person_id
    top5: List[int]      # lista top-5 person_id
    confidence: float    # pewność top-1 (0-1 lub score)
    time_ms: float
    correct: bool        # predicted_id == true_id


@dataclass
class FoldSummary:
    """Wyniki jednego folda CV."""

    fold: int
    method: str
    accuracy: float
    top3_accuracy: float
    top5_accuracy: float
    precision: float
    recall: float
    f1: float
    time_s: float


class ExperimentResults:
    """Kolekcja wyników eksperymentu gotowa do zapisu."""

    def __init__(self, experiment_name: str = "") -> None:
        self.experiment_name = experiment_name or f"experiment_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.timestamp = datetime.now().isoformat(timespec="seconds")
        self.predictions: List[PredictionRow] = []
        self.fold_summaries: List[FoldSummary] = []
        self.confusion_matrices: Dict[str, np.ndarray] = {}  # method → 22×22
        self.confusion_classes: Dict[str, List[int]] = {}    # method → lista klas
        self.run_info: Dict[str, str] = {}

    def add_run_info(self, key: str, value: str) -> None:
        self.run_info[key] = str(value)

    def add_prediction(self, row: PredictionRow) -> None:
        self.predictions.append(row)

    def add_fold_summary(self, summary: FoldSummary) -> None:
        self.fold_summaries.append(summary)

    def add_confusion_matrix(
        self, method: str, matrix: np.ndarray, classes: List[int]
    ) -> None:
        self.confusion_matrices[method] = matrix
        self.confusion_classes[method] = classes

    def save(self, output_path: Optional[Path] = None) -> Path:
        """Zapisz wyniki do pliku Excel.

        Args:
            output_path: ścieżka do pliku .xlsx (domyślnie results/<name>.xlsx).

        Returns:
            Ścieżka do zapisanego pliku.
        """
        if output_path is None:
            RESULTS_DIR.mkdir(parents=True, exist_ok=True)
            output_path = RESULTS_DIR / f"{self.experiment_name}.xlsx"
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Usuń domyślny pusty arkusz

        self._write_run_info(wb)
        if self.predictions:
            self._write_predictions(wb)
        if self.fold_summaries:
            self._write_summaries(wb)
        for method, matrix in self.confusion_matrices.items():
            self._write_confusion_matrix(wb, method, matrix, self.confusion_classes[method])

        wb.save(str(output_path))
        return output_path

    # ----- Prywatne metody zapisu arkuszy -----

    def _header_style(self, cell) -> None:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")

    def _write_run_info(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Run_info")
        ws.append(["Parametr", "Wartość"])
        self._header_style(ws["A1"])
        self._header_style(ws["B1"])

        ws.append(["Timestamp", self.timestamp])
        ws.append(["Experiment", self.experiment_name])
        ws.append(["Łączna liczba predykcji", len(self.predictions)])
        ws.append(["Metody", ", ".join(sorted({p.method for p in self.predictions}))])
        for k, v in self.run_info.items():
            ws.append([k, v])

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

    def _write_predictions(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Predictions")
        headers = [
            "image_path", "true_id", "method", "predicted_id",
            "top3", "top5", "confidence", "time_ms", "correct"
        ]
        ws.append(headers)
        for cell in ws[1]:
            self._header_style(cell)

        for row in self.predictions:
            ws.append([
                row.image_path,
                row.true_id,
                row.method,
                row.predicted_id,
                str(row.top3),
                str(row.top5),
                round(row.confidence, 4),
                round(row.time_ms, 1),
                "TAK" if row.correct else "NIE",
            ])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions["A"].width = 60

    def _write_summaries(self, wb: openpyxl.Workbook) -> None:
        # Pogrupuj po metodzie
        methods: Dict[str, List[FoldSummary]] = {}
        for s in self.fold_summaries:
            methods.setdefault(s.method, []).append(s)

        for method, summaries in methods.items():
            sheet_name = f"Summary_{method[:20]}"
            ws = wb.create_sheet(sheet_name)
            headers = ["Fold", "Accuracy", "Top-3", "Top-5", "Precision", "Recall", "F1", "Czas [s]"]
            ws.append(headers)
            for cell in ws[1]:
                self._header_style(cell)

            for s in sorted(summaries, key=lambda x: x.fold):
                ws.append([
                    s.fold, round(s.accuracy, 4), round(s.top3_accuracy, 4),
                    round(s.top5_accuracy, 4), round(s.precision, 4),
                    round(s.recall, 4), round(s.f1, 4), round(s.time_s, 1),
                ])

            # Wiersz średnia ± std
            if len(summaries) > 1:
                accs = [s.accuracy for s in summaries]
                top5s = [s.top5_accuracy for s in summaries]
                f1s = [s.f1 for s in summaries]
                ws.append([
                    "MEAN ± STD",
                    f"{np.mean(accs):.4f} ± {np.std(accs):.4f}",
                    "",
                    f"{np.mean(top5s):.4f} ± {np.std(top5s):.4f}",
                    "", "",
                    f"{np.mean(f1s):.4f} ± {np.std(f1s):.4f}",
                    "",
                ])
                # Pogrub ostatni wiersz
                last_row = ws.max_row
                for cell in ws[last_row]:
                    cell.font = Font(bold=True)

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

    def _write_confusion_matrix(
        self, wb: openpyxl.Workbook, method: str, matrix: np.ndarray, classes: List[int]
    ) -> None:
        sheet_name = f"CM_{method[:22]}"
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value=f"Confusion Matrix — {method}")
        ws.cell(row=1, column=1).font = Font(bold=True)

        # Nagłówki kolumn
        for j, cls in enumerate(classes):
            ws.cell(row=2, column=j + 2, value=f"Nr_{cls:02d}")
            self._header_style(ws.cell(row=2, column=j + 2))

        # Wiersze
        for i, true_cls in enumerate(classes):
            ws.cell(row=i + 3, column=1, value=f"Nr_{true_cls:02d}")
            ws.cell(row=i + 3, column=1).font = Font(bold=True)
            for j, val in enumerate(matrix[i]):
                cell = ws.cell(row=i + 3, column=j + 2, value=int(val))
                if i == j and val > 0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")  # zielony = dobrze
                elif i != j and val > 0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")  # czerwony = błąd

        for col in range(1, len(classes) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 8

